# -*- coding: utf-8 -*-
"""
驷马合规 · 广告审查助手 — 力邦营养企业定制版 API
企业内用版：无配额限制，聚焦食品/特医行业
"""
import json, os, time, base64, io, re, threading, logging, uuid
from typing import Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

from fastapi import FastAPI, HTTPException, Request, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from pydantic import BaseModel, Field

from llm import analyze as llm_analyze, validate_config as llm_validate_config
from detector import detect as detector_detect, compile_kb
from pdf_report import generate_pdf

from user import (
    check_quota, record_use, get_user,
    record_check, get_history, get_record_detail, delete_record, clear_history,
    init as init_db,
    _cleanup_old_records, _get_all_users, _get_admin_stats,
    PLANS
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("main")

ROOT_DIR = os.path.dirname(__file__)
KB = os.path.join(ROOT_DIR, "..", "..", "ad-compliance-checker", "knowledge", "forbidden_words.json")
UPGRADE_SECRET = os.environ.get("ADMIN_SECRET", "")
MAX_TEXT_LENGTH = int(os.environ.get("MAX_TEXT_LENGTH", "5000"))

_ocr_reader = None
_ocr_lock = threading.Lock()
_admin_failures = {}
_ADMIN_FAIL_LIMIT = 10
_ADMIN_FAIL_WINDOW = 300

ENTERPRISE_OPENID = "libang_enterprise"

# ========== 简易频率限制 ==========
_rate_limit_map = {}
def _check_rate_limit(key: str, max_req: int = 30, window: int = 60) -> bool:
    now = time.time()
    info = _rate_limit_map.get(key)
    if not info or now - info[1] > window:
        _rate_limit_map[key] = [1, now]
        return False
    info[0] += 1
    return info[0] > max_req

STATIC_DIRS = {
    "landing": os.path.join(ROOT_DIR, ".."),
    "mobile": os.path.join(ROOT_DIR, "..", "h5"),
    "desktop": os.path.join(ROOT_DIR, "..", "frontend"),
    "batch": os.path.join(ROOT_DIR, "..", "batch"),
    "admin": os.path.join(ROOT_DIR, "admin"),
}

app = FastAPI(title="驷马合规 · 广告审查助手 — 力邦营养企业定制版", version="1.0.0")

_cors_default = "http://127.0.0.1:8000,http://localhost:8000"
app.add_middleware(CORSMiddleware,
    allow_origins=os.environ.get("CORS_ORIGINS", _cors_default).split(","),
    allow_methods=["*"], allow_headers=["*"])
app.add_middleware(GZipMiddleware, minimum_size=1000)


# ========== 统一错误处理 ==========
class APIError(Exception):
    def __init__(self, message: str, status_code: int = 200):
        self.message = message
        self.status_code = status_code

@app.exception_handler(APIError)
async def api_error_handler(request: Request, exc: APIError):
    return JSONResponse(status_code=exc.status_code, content={"ok": False, "error": exc.message})

@app.exception_handler(HTTPException)
async def http_error_handler(request: Request, exc: HTTPException):
    return JSONResponse(status_code=exc.status_code, content={"ok": False, "error": exc.detail})

@app.middleware("http")
async def add_error_handler(request: Request, call_next):
    try:
        return await call_next(request)
    except HTTPException as exc:
        return JSONResponse(status_code=exc.status_code, content={"ok": False, "error": exc.detail})
    except Exception as exc:
        logger.exception("未处理的异常")
        return JSONResponse(status_code=500, content={"ok": False, "error": "服务器内部错误"})


# ========== 请求模型 ==========
class CheckReq(BaseModel):
    text: str
    industry: str = ""
    platform: str = ""
    openid: str = ENTERPRISE_OPENID

class BatchCheckReq(BaseModel):
    texts: list[str] = []
    text: str = ""
    industry: str = ""
    platform: str = ""
    openid: str = ENTERPRISE_OPENID

class PaginationReq(BaseModel):
    openid: str = ENTERPRISE_OPENID
    page: int = Field(default=1, ge=1)
    page_size: int = Field(default=20, ge=1, le=100)

class RecordReq(BaseModel):
    openid: str = ENTERPRISE_OPENID
    record_id: int

class ClearReq(BaseModel):
    openid: str = ENTERPRISE_OPENID


def _validate_text(text: str) -> str:
    if not text or not text.strip():
        raise HTTPException(status_code=400, detail="文案内容不能为空")
    text = text.strip()
    if len(text) > MAX_TEXT_LENGTH:
        raise HTTPException(status_code=400, detail=f"文案超出长度限制（最多{MAX_TEXT_LENGTH}字）")
    return text

def _get_llm_api_key() -> str:
    for key in ["MIMO_API_KEY", "OPENAI_API_KEY", "DEEPSEEK_API_KEY"]:
        val = os.environ.get(key, "")
        if val:
            return val
    return ""


@app.on_event("startup")
def on_startup():
    init_db()
    compile_kb()
    for path in STATIC_DIRS.values():
        os.makedirs(path, exist_ok=True)
    try:
        _cleanup_old_records()
    except Exception:
        logger.warning("清理旧记录失败（不影响启动）")
    llm_ok, llm_msg = llm_validate_config()
    if llm_ok:
        logger.info(f"LLM 配置校验通过: {llm_msg}")
    else:
        logger.warning(f"LLM 配置未通过: {llm_msg}；AI 深度分析不可用")
    logger.info("力邦营养企业定制版 v1.0 已启动，无配额限制")


# ========== 核心检测接口 ==========
@app.post("/api/check")
async def check(req: CheckReq):
    text = _validate_text(req.text)
    result = detector_detect(text, req.industry, req.platform)
    record_use(req.openid)
    record_check(req.openid, text, req.industry, "quick", result)
    return {"ok": True, **result}


@app.post("/api/deep-check")
async def deep_check(req: CheckReq):
    text = _validate_text(req.text)
    if not _get_llm_api_key():
        raise APIError("AI 深度分析暂不可用（未配置 API Key）")
    result = llm_analyze(text, req.industry)
    record_use(req.openid, deep=True)
    record_check(req.openid, text, req.industry, "deep",
                 {"score": 0, "risk_count": 0, "summary": "AI deep analysis complete", "risks": []})
    return {"ok": True, "analysis": result}


@app.post("/api/batch-check")
async def batch_check(req: BatchCheckReq, request: Request = None):
    if request:
        client_ip = request.client.host if request.client else "unknown"
        rate_key = f"batch:{client_ip}"
        if _check_rate_limit(rate_key, 30, 60):
            raise APIError("请求过于频繁，请稍后再试")

    texts = req.texts
    if not texts:
        lines = req.text.strip().split("\n")
        if any("---" in l.strip() for l in lines if l.strip()):
            texts = [t.strip() for t in re.split(r"\n\s*---\s*\n|\n\s*---\s*$|^\s*---\s*\n", req.text.strip()) if t.strip()]
        if not texts or len(texts) <= 1:
            texts = [t.strip() for t in lines if t.strip() and "---" not in t.strip()]
    if not texts:
        raise APIError("请输入至少一条文案")
    if len(texts) > 20:
        raise APIError(f"批量检测最多支持 20 条文案（当前 {len(texts)} 条）")

    results = []
    for i, t in enumerate(texts):
        if len(t) > 5000:
            results.append({"index": i, "text_preview": t[:50], "error": "文案超长"})
            continue
        try:
            r = detector_detect(t, req.industry, req.platform)
            results.append({
                "index": i, "text_preview": t[:80] + ("..." if len(t) > 80 else ""),
                "score": r["score"], "risk_count": r["risk_count"],
                "risks": r["risks"][:5], "summary": r["summary"],
            })
            record_check(req.openid, t, req.industry, "batch", r)
        except Exception as e:
            results.append({"index": i, "text_preview": t[:50], "error": str(e)})

    record_use(req.openid)
    return {"ok": True, "results": results, "total": len(texts), "success": sum(1 for r in results if "error" not in r)}


@app.post("/api/export-pdf")
async def export_pdf(req: CheckReq):
    text = _validate_text(req.text)
    result = detector_detect(text, req.industry, req.platform)
    record_check(req.openid, text, req.industry, "pdf", result)
    raw = generate_pdf(text, result)
    if isinstance(raw, bytearray):
        raw = bytes(raw)
    elif isinstance(raw, memoryview):
        raw = bytes(raw)
    is_pdf = raw[:4] == b"%PDF"
    return Response(
        content=raw,
        media_type="application/pdf" if is_pdf else "text/plain; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="compliance_report_{int(time.time())}.{"pdf" if is_pdf else "txt"}"'
        }
    )


# ========== 用户与历史 ==========
@app.get("/api/user")
async def user_info(openid: str = ENTERPRISE_OPENID):
    return {"ok": True, "data": get_user(openid)}

@app.post("/api/history")
async def history(req: PaginationReq):
    return {"ok": True, "data": get_history(req.openid, req.page, req.page_size)}

@app.post("/api/record")
async def record_detail(req: RecordReq):
    detail = get_record_detail(req.record_id, req.openid)
    if not detail:
        raise APIError("记录不存在")
    return {"ok": True, "data": detail}

@app.post("/api/delete-record")
async def delete_record_api(req: RecordReq):
    ok = delete_record(req.record_id, req.openid)
    return {"ok": ok}

@app.post("/api/clear-history")
async def clear_history_api(req: ClearReq):
    count = clear_history(req.openid)
    return {"ok": True, "deleted": count}


# ========== OCR ==========
@app.post("/api/ocr-base64")
async def ocr_base64(data: dict):
    try:
        import easyocr
        import numpy as np
        from PIL import Image
    except ImportError:
        return {"text": "", "ok": False, "error": "easyocr not installed"}

    global _ocr_reader
    if _ocr_reader is None:
        with _ocr_lock:
            if _ocr_reader is None:
                try:
                    _ocr_reader = easyocr.Reader(["ch_sim", "en"], gpu=False)
                except Exception as e:
                    return {"text": "", "ok": False, "error": f"OCR 模型加载失败: {e}"}

    with _ocr_lock:
        try:
            img_data = data.get("image", "")
            if "," in img_data:
                img_data = img_data.split(",")[1]
            img_bytes = base64.b64decode(img_data)
            img = Image.open(io.BytesIO(img_bytes))
            MAX_SIDE = 1600
            if img.width > MAX_SIDE or img.height > MAX_SIDE:
                ratio = MAX_SIDE / max(img.width, img.height)
                img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
            results = _ocr_reader.readtext(np.array(img), detail=0)
            text = "\n".join(results) if results else ""
            return {"text": text.strip(), "ok": True}
        except Exception as e:
            return {"text": "", "ok": False, "error": str(e)}


# ========== 批量OCR图片审查 ==========
_BATCH_RESULTS = {}  # {batch_id: {results: [...], status: str}}

def _ocr_single_image(img_bytes: bytes) -> str:
    """对单张图片执行OCR，返回文字"""
    global _ocr_reader
    if _ocr_reader is None:
        with _ocr_lock:
            if _ocr_reader is None:
                try:
                    import easyocr
                    _ocr_reader = easyocr.Reader(["ch_sim", "en"], gpu=False)
                except Exception as e:
                    return f"[OCR加载失败: {e}]"
    try:
        from PIL import Image
        import numpy as np
        img = Image.open(io.BytesIO(img_bytes))
        MAX_SIDE = 1600
        if img.width > MAX_SIDE or img.height > MAX_SIDE:
            ratio = MAX_SIDE / max(img.width, img.height)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
        with _ocr_lock:
            results = _ocr_reader.readtext(np.array(img), detail=0)
            return "\n".join(results) if results else ""
    except Exception as e:
        return f"[OCR错误: {e}]"


@app.post("/api/batch-ocr-check")
async def batch_ocr_check(files: list[UploadFile] = File(...), industry: str = Form(""), platform: str = Form("")):
    """批量上传图片 → OCR → 合规检测"""
    if not files:
        raise APIError("请上传至少一张图片")
    if len(files) > 50:
        raise APIError(f"批量最多支持50张图片（当前{len(files)}张）")

    batch_id = uuid.uuid4().hex[:12]
    total = len(files)
    results = [None] * total

    def process_one(i: int, f: UploadFile) -> tuple:
        """处理单张图片"""
        name = getattr(f, "filename", f"图片{i+1}") or f"图片{i+1}"
        try:
            img_bytes = f.file.read()
            # OCR提取文字
            text = _ocr_single_image(img_bytes)
            if not text or text.startswith("[OCR"):
                return i, {"index": i, "name": name, "error": text if text else "OCR未识别到文字", "status": "ocr_failed"}
            # 合规检测
            r = detector_detect(text, industry, platform)
            risks = r.get("risks", [])
            forbidden = [x for x in risks if x.get("type") in ("forbidden", "industry")]
            risk_words = [x["word"] for x in forbidden[:10]]
            return i, {
                "index": i,
                "name": name,
                "text_preview": text[:150] + ("..." if len(text) > 150 else ""),
                "text_full": text[:2000],
                "score": r["score"],
                "risk_count": r["risk_count"],
                "risk_words": risk_words,
                "summary": r["summary"],
                "status": "ok",
                "compliance_tips": r.get("compliance_tips", []),
            }
        except Exception as e:
            return i, {"index": i, "name": name, "error": str(e), "status": "error"}

    # 并行处理（最多5个并发）
    pool_size = min(5, total)
    with ThreadPoolExecutor(max_workers=pool_size) as executor:
        futures = {executor.submit(process_one, i, f): i for i, f in enumerate(files)}
        for future in as_completed(futures):
            i, result = future.result()
            results[i] = result

    success = sum(1 for r in results if r and r.get("status") == "ok")
    _BATCH_RESULTS[batch_id] = {"results": results, "total": total, "success": success, "created": time.time()}

    return {
        "ok": True,
        "batch_id": batch_id,
        "total": total,
        "success": success,
        "results": results,
    }


@app.get("/api/batch-results/{batch_id}")
async def get_batch_results(batch_id: str):
    """获取批量检测结果（供导出使用）"""
    data = _BATCH_RESULTS.get(batch_id)
    if not data:
        raise APIError("批次不存在或已过期")
    return {"ok": True, "data": data}


# 清理30分钟前的历史批次
def _cleanup_old_batches():
    now = time.time()
    expired = [k for k, v in _BATCH_RESULTS.items() if now - v.get("created", 0) > 1800]
    for k in expired:
        _BATCH_RESULTS.pop(k, None)


_cleanup_thread = threading.Thread(target=_cleanup_old_batches, daemon=True)
_cleanup_thread.start()


# ========== 批量合规报告导出（Excel）==========
@app.post("/api/batch-export-excel")
async def batch_export_excel(req: dict):
    """导出批量检测结果Excel报告"""
    batch_id = req.get("batch_id", "")
    data = _BATCH_RESULTS.get(batch_id) if batch_id else None
    if not data:
        # 如果批次过期或没有batch_id，直接从results参数构建
        results = req.get("results", [])
        if not results:
            raise APIError("无检测结果可供导出")
        data = {"results": results, "total": len(results), "success": sum(1 for r in results if r.get("status") == "ok")}

    try:
        excel_bytes = _generate_batch_excel(data)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="力邦营养_批量合规报告_{int(time.time())}.xlsx"'
            }
        )
    except ImportError:
        raise APIError("Excel导出需要安装 openpyxl：pip install openpyxl")
    except Exception as e:
        logger.exception("Excel导出失败")
        raise APIError(f"导出失败: {e}")


def _generate_batch_excel(data: dict) -> bytes:
    """生成批量合规报告 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "批量合规报告"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 标题行
    ws.merge_cells("A1:H1")
    ws["A1"] = "驷马合规 · 力邦营养企业定制版 — 批量广告合规检测报告"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2563EB")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"生成时间: {time.strftime('%Y-%m-%d %H:%M')}　|　总计: {data['total']}条　|　检测完成: {data['success']}条"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 24

    # 表头
    headers = ["序号", "图片/产品名称", "识别文字摘要", "得分", "风险数", "违规词", "检测摘要", "状态"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    # 设置列宽
    col_widths = [6, 20, 40, 8, 8, 25, 30, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 风险色
    red_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    green_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")

    # 数据行
    for idx, r in enumerate(data.get("results", [])):
        row = idx + 4
        status = r.get("status", "ok")
        score = r.get("score", 0) if status == "ok" else 0

        fill = None
        if status == "ok":
            if score < 60:
                fill = red_fill
            elif score < 80:
                fill = yellow_fill
            else:
                fill = green_fill
        else:
            fill = red_fill

        values = [
            idx + 1,
            r.get("name", ""),
            r.get("text_preview", r.get("text_full", "")[:150]) if status == "ok" else r.get("error", "错误"),
            score if status == "ok" else "-",
            r.get("risk_count", 0) if status == "ok" else "-",
            "、".join(r.get("risk_words", [])[:8]) if status == "ok" else "",
            r.get("summary", "") if status == "ok" else "检测失败",
            "✓ 正常" if status == "ok" else "✗ 失败",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="微软雅黑", size=10)
            if fill:
                cell.fill = fill

        ws.row_dimensions[row].height = 32

    return wb.save_to_memory() if hasattr(wb, "save_to_memory") else _wb_to_bytes(wb)


def _wb_to_bytes(wb) -> bytes:
    """Workbook 转 bytes（openpyxl 兼容）"""
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ========== 信息 ==========
@app.get("/api/info")
async def api_info():
    return {
        "name": "驷马合规 · 广告审查助手（力邦营养企业定制版）",
        "version": "1.0.0",
        "enterprise": "力邦营养",
        "endpoints": ["/api/health", "/api/info", "/api/user", "/api/check", "/api/deep-check", "/api/history", "/api/record"]
    }

@app.get("/api/plans")
async def get_plans():
    return {"plans": [{"id": "enterprise", "name": "企业版", "quick": 999999, "deep": 999999}]}

@app.get("/api/health")
async def health():
    return {"status": "ok", "timestamp": time.time()}


# ========== 静态文件 ==========
@app.get("/")
async def serve_index():
    index_path = os.path.join(STATIC_DIRS["landing"], "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path, media_type="text/html")
    return RedirectResponse(url="/m/")

app.mount("/m", StaticFiles(directory=STATIC_DIRS["mobile"], html=True), name="mobile")
app.mount("/pc", StaticFiles(directory=STATIC_DIRS["desktop"], html=True), name="desktop")
app.mount("/batch", StaticFiles(directory=STATIC_DIRS["batch"], html=True), name="batch")

# 管理后台
_ADMIN_DIR = STATIC_DIRS["admin"]
_ADMIN_LOGIN = os.path.join(_ADMIN_DIR, "login.html")
_ADMIN_INDEX = os.path.join(_ADMIN_DIR, "index.html")

@app.get("/admin/login.html")
async def admin_login_page():
    return FileResponse(_ADMIN_LOGIN, media_type="text/html")

@app.get("/admin/index.html")
async def admin_index_page(request: Request):
    if UPGRADE_SECRET and request.cookies.get("admin_ok") != "1":
        return RedirectResponse(url="/admin/login.html")
    return FileResponse(_ADMIN_INDEX, media_type="text/html")

@app.get("/admin")
async def admin_redirect():
    return RedirectResponse(url="/admin/index.html")

app.mount("/static", StaticFiles(directory=STATIC_DIRS["mobile"]), name="static")
